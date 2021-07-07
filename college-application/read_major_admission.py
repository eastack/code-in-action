#!/usr/bin/env python
import psycopg2
from openpyxl import load_workbook

wb = load_workbook('/home/radix10/downloads/college-result.xlsx')
ws = wb.active

host = "localhost"
dbname = "volu"
user = "eastack"
password = "eastack"

conn_string = f"host={host} user={user} dbname={dbname} password={password}"
conn = psycopg2.connect(conn_string)
cursor = conn.cursor()

for row in ws.iter_rows(min_row=4, min_col=2, max_col=8, max_row=16717):
    major_info = row[0].value.strip()
    college_info = row[1].value.strip()

    cursor.execute(f'insert into college_result(college_name, college_code, major_name, major_code, lower_score, lower_rank) '
                   f"values('{college_info[4:]}', '{college_info[0:4]}', '{major_info[2:]}', '{major_info[0:2]}', {row[5].value}, {row[6].value})")

conn.commit()
cursor.close()
conn.close()