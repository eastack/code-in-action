#!/usr/bin/env python
import psycopg2
from openpyxl import Workbook

host = "47.95.120.5"
dbname = "college_application"
user = "mn"
password = "chrdwhdhxt"

conn_string = f"host={host} user={user} dbname={dbname} password={password}"
conn = psycopg2.connect(conn_string)
cursor = conn.cursor()

applications = [
    'E100:01',
    'A430:30',
    'A434:12',
    'A447:1R',
    'A434:32',
    'A447:1P',
    'E100:10',
    'A430:23',
    'A439:40',
    'E100:02',
    'A435:60',
    'A452:40',
    'A439:46',
    'A434:81',
    'A439:33',
    'A435:0U',
    'B688:13',
    'B510:11',
    'A452:15',
    'A447:22',
    'A447:25',
    'A438:02',
    'A443:24',
    'A448:0T',
    'A434:33',
    'B688:38',
    'A443:23',
    'B067:57',
    'A904:0Q',
    'B067:27',
    'A449:48',
    'A434:02',
    'B510:18',
    'A438:33',
    'A443:30',
    'A449:45',
    'A455:44',
    'A455:50',
    'E227:36',
    'B067:54',
    'A449:36',
    'A452:15',
    'A447:0J',
    'E438:19',
    'A455:24',
    'A453:1Q',
    'A454:38',
    'C331:21',
    'A443:24',
    'A439:54',
    'B067:20',
    'A454:35',
    'A448:34',
    'E439:25',
    'A454:41',
    'A904:0R',
    'A454:24',
    'A443:29',
    'A455:43',
    'A453:28',
    'A440:1T',
    'A447:1V',
    'A433:21',
    'A430:49',
    'A441:16',
    'A443:89',
    'A443:88',
    'A431:1V',
    'E276:37',
    'E276:49',
    'A433:64',
    'A441:05',
    'A443:64',
    'E277:04',
    'E277:07',
    'B653:41',
    'B067:14',
    'A483:54',
    'A482:12',
    'A483:7E',
    'A465:37',
    'A478:44',
    'A483:05',
    'A483:35',
    'A478:35',
    'A465:09',
    'A478:11',
    'A478:14',
    'A478:12',
    'A479:07',
    'A467:21',
    'A483:56',
    'B517:0Y',
    'A071:27',
    'D658:03',
    'D658:12',
]

wb = Workbook()
ws = wb.active
ws.title = "志愿表"

for row, app in enumerate(applications, start=1):
    a = app.split(":")
    college_code = a[0]
    major_code = a[1]
    cursor.execute(
        f"select major_name, college_name, lower_score from college_result where college_code = '{college_code}' and major_code = '{major_code}'")
    rows = cursor.fetchall()

    if (len(rows) > 1):
        print(rows)
        conn.commit()
        cursor.close()
        conn.close()
        exit(1)

    if (len(rows) == 0):
        print(f"{college_code}:{major_code} 错误组合不存在!!!")
    else:
        major_name = rows[0][0]
        college_name = rows[0][1]
        lower_score = rows[0][2]
        cursor.execute(
            f"select major_code from major_admission_tj where school_code = '{college_code}' and major_name = '{major_name}'")

        rows = cursor.fetchall()

        ws.cell(row=row, column=1, value=lower_score)
        ws.cell(row=row, column=2, value=college_name)
        ws.cell(row=row, column=3, value=college_code)
        ws.cell(row=row, column=4, value=major_name)
        ws.cell(row=row, column=5, value=major_code)
        ws.cell(row=row, column=6, value=str(rows))

    print(f"{lower_score}::{college_code}/{college_name}::{major_code}/{major_name}-->{rows}")

wb.save('balances-tj.xlsx')
conn.commit()
cursor.close()
conn.close()
